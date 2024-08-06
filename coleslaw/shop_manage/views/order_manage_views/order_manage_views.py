from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.views.generic import View
from django.http import HttpRequest, JsonResponse
from django.http.response import HttpResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger, InvalidPage
from django.db.models import CharField, F, Value as V, Func, Sum, Case, When, IntegerField
from django.db.models.functions import Coalesce, Cast
from django.utils.decorators import method_decorator
from django.db import transaction
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from system_manage.decorators import permission_required
from system_manage.models import Order, OrderGoods, ShopAdmin, OrderPayment
from system_manage.utils import ResponseToXlsx
from shop_manage.views.shop_manage_views.auth_views import check_shop
from api.views.sms_views.sms_views import send_sms

from openpyxl import Workbook
from openpyxl.styles import Font
import json, datetime
import urllib.parse

class OrderManageView(View):
    '''
        주문내역 관리 화면
    '''
    @method_decorator(permission_required(redirect_url='shop_manage:denied'))
    def get(self, request: HttpRequest, *args, **kwargs):
        context = {}
        shop_id = kwargs.get('shop_id')
        shop = check_shop(pk=shop_id)
        if not shop:
            return redirect('shop_manage:notfound')
        context['shop'] = shop
        
        paginate_by = '20'
        page = request.GET.get('page', '1')
        search_type = request.GET.get('search_type', '')
        search_keyword = request.GET.get('search_keyword', '')
        condition = request.GET.get('condition', '')

        filter_dict = {}
        filter_dict['shop'] = shop

        if condition:
            order_date_no = request.GET.get('order_date_no', '0') 
            dates = request.GET.get('dates', '')
            context['dates'] = dates
            if dates != '':
                startDate = dates.split(' - ')[0].strip()
                endDate = dates.split(' - ')[1].strip()
                format = '%m/%d/%Y'

                startDate = datetime.datetime.strptime(startDate, format)
                endDate = datetime.datetime.strptime(endDate, format)
                endDate = datetime.datetime.combine(endDate, datetime.time.max)
                filter_dict['created_at__lte'] = endDate
                filter_dict['created_at__gte'] = startDate
        else:
            order_date_no = '1'
            today = timezone.now().strftime("%m/%d/%Y")
            context['dates'] = f"{today} - {today}"
            filter_dict['date'] = timezone.now().date()
            
        status_list = ['1', '2', '3', '4', '5']
        filter_dict['status__in'] = status_list
        context['order_date_no'] = order_date_no
        
        if search_keyword:
            context['search_type'] = search_type
            context['search_keyword'] = search_keyword
            filter_dict[search_type + '__icontains'] = search_keyword

        aggregate = request.GET.get('aggregate', None)
        if aggregate:
            new_filter_dict = {'order__' + str(key): val for key, val in filter_dict.items()}
            filename = f"{shop.name_kr} 판매내역"
            headers = ['대분류', '소분류', '상품명', '옵션명', '상품가격', '옵션가격', '총가격', '수량']
            queryset = OrderGoods.objects.filter(**new_filter_dict)

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename*=UTF-8\'\'%s.xlsx' % urllib.parse.quote(filename.encode('utf-8'))

            wb = Workbook()
            ws = wb.active
            ws.title = "판매완료 현황"
            # Add headers
            ws.append(headers)
            # Add data from the model
            complete_queryset = queryset.filter(order__status__in=['1', '3', '4', '5']).annotate(sale_total_price=F('sale_price') + F('sale_option_price')).values("goods__sub_category__main_category__name_kr", "goods__sub_category__name_kr", "name_kr", "option_kr", "sale_price", "sale_option_price", "sale_total_price").annotate(total_quantity=Sum("quantity")).order_by('goods__sub_category__main_category__name_kr', 'goods__sub_category__name_kr', 'name_kr', '-total_quantity')
            for i in complete_queryset:
                ws.append([i["goods__sub_category__main_category__name_kr"], i["goods__sub_category__name_kr"], i['name_kr'], i['option_kr'], i['sale_price'], i['sale_option_price'], i['sale_total_price'], i['total_quantity']])

            ws1 = wb.create_sheet('취소 현황')
            ws1.append(headers)
            cancel_queryset = queryset.filter(order__status='2').annotate(sale_total_price=F('sale_price') + F('sale_option_price')).values("goods__sub_category__main_category__name_kr", "goods__sub_category__name_kr", "name_kr", "option_kr", "sale_price", "sale_option_price", "sale_total_price").annotate(total_quantity=Sum("quantity")).order_by('goods__sub_category__main_category__name_kr', 'goods__sub_category__name_kr', 'name_kr', '-total_quantity')
            for i in cancel_queryset:
                ws1.append([i["goods__sub_category__main_category__name_kr"], i["goods__sub_category__name_kr"], i['name_kr'], i['option_kr'], i['sale_price'], i['sale_option_price'], i['sale_total_price'], i['total_quantity']])

            ws2 = wb.create_sheet('부분 취소 현황')
            ws2.append(headers)
            cancel_queryset2 = queryset.filter(order__status='6').annotate(sale_total_price=F('sale_price') + F('sale_option_price')).values("goods__sub_category__main_category__name_kr", "goods__sub_category__name_kr", "name_kr", "option_kr", "sale_price", "sale_option_price", "sale_total_price").annotate(total_quantity=Sum("quantity")).order_by('goods__sub_category__main_category__name_kr', 'goods__sub_category__name_kr', 'name_kr', '-total_quantity')
            for i in cancel_queryset2:
                ws2.append([i["goods__sub_category__main_category__name_kr"], i["goods__sub_category__name_kr"], i['name_kr'], i['option_kr'], i['sale_price'], i['sale_option_price'], i['sale_total_price'], i['total_quantity']])

            # Save the workbook to the HttpResponse
            wb.save(response)
            return response

        excel = request.GET.get('excel', None)
        if excel:
            new_filter_dict = {'order__' + str(key): val for key, val in filter_dict.items()}
            filename = f"{shop.name_kr} 결제내역"
            columns = ['주문 ID', '주문번호(QR)', '거래번호(QR)', '승인번호(QR)', '승인번호(POS)', '승인날짜', '승인시간', '날짜', '발급사명', '카드정보', '주문정보', '부가세', '결제금액', '상태', '주문타입']
            queryset = OrderPayment.objects.filter(**new_filter_dict).annotate(
                signedAmount=Case(
                    When(order__status='2', then= Cast(F('amount'), IntegerField()) * 0),
                    default=F('amount'), output_field=IntegerField()
                ),
                paymentStatus=Case(
                    When(status=True, then=V('결제완료')),
                    When(status=False, then=V('결제취소')),
                ),
                orderType=Case(
                    When(order__order_type='0', then=V('POS')),
                    When(order__order_type='1', then=V('QR')),
                    When(order__order_type='2', then=V('KIOSK'))
                ),     
            ).values(
                'order_id',
                'mbrRefNo',
                'refNo',
                'applNo',
                'approvalNumber',
                'tranDate',
                'tranTime',
                'created_at',
                'issueCompanyName',
                'cardNo',
                'order__order_name_kr',
                'taxAmount',
                'signedAmount',
                'paymentStatus',
                'orderType'
            ).order_by('-id')
            xlsx_download = ResponseToXlsx(columns=columns, queryset=queryset)
            return xlsx_download.download(filename=filename)
            

        obj_list = Order.objects.filter(**filter_dict).annotate(
        createdAt = Func(
            F('created_at'),
            V('%y.%m.%d %H:%i'),
            function='DATE_FORMAT',
            output_field=CharField()
        )).order_by('-created_at').values(
            'id',
            'order_no',
            'order_name_kr',
            'order_name_en',
            'order_code',
            'order_membername',
            'order_phone',
            'final_price',
            'status',
            'order_complete_sms',
            'createdAt'
        )

        total_price = obj_list.exclude(status='2').aggregate(sum=Coalesce(Sum('final_price'), 0)).get('sum')
        context['total_price'] = total_price

        paginator = Paginator(obj_list, paginate_by)
        try:
            page_obj = paginator.page(page)
        except PageNotAnInteger:
            page = 1
            page_obj = paginator.page(page)
        except EmptyPage:
            page = 1
            page_obj = paginator.page(page)
        except InvalidPage:
            page = 1
            page_obj = paginator.page(page)

        pagelist = paginator.get_elided_page_range(page, on_each_side=3, on_ends=1)
        context['pagelist'] = pagelist
        context['page_obj'] = page_obj

        return render(request, 'order_manage/order_manage.html', context)
    
    @method_decorator(permission_required(raise_exception=True))
    def put(self, request: HttpRequest, *args, **kwargs):
        request.PUT = json.loads(request.body)
        order_id = request.PUT['order_id']
        order_status = request.PUT['order_status']
        status = ['1', '3', '4', '5']
        try:
            order = Order.objects.get(pk=order_id)
        except:
            return JsonResponse({'message' : '데이터 오류'},  status = 400)
        if order_status not in status:
            return JsonResponse({'message' : '상태 값 오류'},  status = 400)
        order.status = order_status
        order.save()
        
        return JsonResponse({'message' : '수정되었습니다.'},status = 200)


@require_http_methods(["POST"])
def order_complete_sms(request: HttpRequest, *args, **kwargs):
    '''
        주문완료 문자
    '''
    shop_id = kwargs.get('shop_id')
    shop = check_shop(pk=shop_id)
    if not shop:
        return JsonResponse({"message":"잘못된 가맹점입니다."}, status = 400)
    order_id = request.POST['order_id']
    try:
        order = Order.objects.get(pk=order_id, shop=shop)
    except:
        return JsonResponse({"message":"잘못된 주문입니다."}, status = 400)
    if request.user.is_superuser or ShopAdmin.objects.filter(shop=shop, user=request.user).exists():
        pass
    else:
        return JsonResponse({"message":"권한이 없습니다."}, status = 400)    
    if order.order_complete_sms:
        return JsonResponse({"message":"이미 문자 발송 처리된 주문입니다."}, status = 400)
    if not order.order_phone:
        return JsonResponse({"message":"연락처가 없습니다."}, status = 400)
    
    message=f'[{shop.name_kr}]\n주문번호 [{order.order_no}] 회원님 주문하신거 수령하세요~\n'
    sms_response = send_sms(phone=order.order_phone, message=message)
    if sms_response.status_code != 202:
        return JsonResponse({"message":"전송실패.."}, status = 200)
    
    order.order_complete_sms = True
    order.save()

    return JsonResponse({"message":"전송되었습니다."}, status = 200)


@require_http_methods(["GET"])
def order_goods(request: HttpRequest, *args, **kwargs):
    '''
        주문상세
    '''
    shop_id = kwargs.get('shop_id')
    order_id = kwargs.get('order_id')
    shop = check_shop(pk=shop_id)
    if not shop:
        return JsonResponse({"message":"잘못된 가맹점입니다."}, status = 400)
    try:
        order = Order.objects.get(pk=order_id, shop=shop)
    except:
        return JsonResponse({"message":"잘못된 주문입니다."}, status = 400)
    
    data = {}

    data['id'] = order.pk
    data['createdAt'] = order.created_at.strftime('%Y년 %m월 %d일 %H:%M')
    data['order_code'] = order.order_code
    data['order_no'] = order.order_no
    data['final_price'] = order.final_price
    data['status'] = order.status
    data['order_complete_sms'] = order.order_complete_sms

    order_goods = order.order_goods.all().values( 
        'name_kr',
        'price',
        'option_kr',
        'option_price',
        'quantity',
        'total_price'
    ).order_by('id')
    data['order_goods'] = list(order_goods)

    return JsonResponse(data=data, status = 200)
